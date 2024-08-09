from openpyxl import Workbook
from openpyxl import load_workbook
from random import randint
from datetime import date
import os
import time
import string
import random
from shutil import copy
import GermanyStrings
import pysftp
# Randomly generate a new global ID
lengthOfGlobalID = 9
globalID = ''.join(["{}".format(randint(0, 9)) for num in range(0, lengthOfGlobalID)])

# Create a location to save created files
directory = "Germany Global ID " + globalID
path = os.path.join(os.getcwd(), directory)
os.mkdir(path)

# Randomize RR email
randomString = ''.join(random.choice(string.ascii_uppercase + string.digits) for _ in range(8))
RRemail = GermanyStrings.RRFirstName + GermanyStrings.RRLastName + str(globalID)[5:] + "@bms.com"
#RRemail = "DEUDemo@protonmail.com"
# Get timestamp for file naming
fdate = date.today().strftime('%m%d%Y')
timestamp = time.strftime("%H%M")
MMDDYYYYHHMM = str(fdate) + timestamp
actorProfiles = "Actor_Profile_Test_" + MMDDYYYYHHMM + ".xlsx"
REMSRoster = "Roster_Test_" + MMDDYYYYHHMM + ".xlsx"
siteMasterData = "SITE_MASTER_TEST_REMS_" + MMDDYYYYHHMM + ".xlsx"

# Create the Site Master Data Sheet
siteMasterDataworkbook = Workbook()
sheet = siteMasterDataworkbook.active
sheet["A1"] = "Global_ID"
sheet["B1"] = "AccountName"
sheet["C1"] = "AccountType"
sheet["D1"] = "Product_Activation"
sheet["E1"] = "DEA"
sheet["F1"] = "AddressLine1"
sheet["G1"] = "City"
sheet["H1"] = "State"
sheet["I1"] = "ZIP"
sheet["J1"] = "Country"
sheet["K1"] = "AffilEntityID"
sheet["L1"] = "AffilEntityName"
sheet["M1"] = "Affil_AccountType"
sheet["N1"] = "Affil_DEA"
sheet["O1"] = "Affil_AddressLine1"
sheet["P1"] = "Affil_City"
sheet["Q1"] = "Affil_State"
sheet["R1"] = "Affil_ZIP"
sheet["S1"] = "Affil_Country"
sheet["T1"] = "AffilEntityType"

sheet["A2"] = globalID
sheet["B2"] = GermanyStrings.accountName
sheet["C2"] = "Hospital"
sheet["D2"] = GermanyStrings.product
sheet["F2"] = GermanyStrings.siteAddress
sheet["G2"] = GermanyStrings.siteCity
sheet["I2"] = GermanyStrings.siteZip
sheet["J2"] = GermanyStrings.country
siteMasterDataworkbook.save(filename=siteMasterData)
siteMasterDataworkbook.close()

copy(siteMasterData, path)

# actor profiles sheet
actorProfilesWorkbook = Workbook()
sheet = actorProfilesWorkbook.active
sheet["A1"] = "Role"
sheet["B1"] = "First Name"
sheet["C1"] = "Last Name"
sheet["D1"] = "Job Title"
sheet["E1"] = "Global_ID"
sheet["F1"] = "Account Name"
sheet["G1"] = "Credentials"
sheet["H1"] = "Primary Phone"
sheet["I1"] = "Fax"
sheet["J1"] = "Email Address"
sheet["K1"] = "Product_Activation"

sheet["A2"] = "Authorized Representative"
sheet["B2"] = GermanyStrings.RRFirstName
sheet["C2"] = GermanyStrings.RRLastName
sheet["D2"] = "Authorized Representative"
sheet["E2"] = globalID
sheet["F2"] = GermanyStrings.accountName
sheet["G2"] = GermanyStrings.RRCredentials
sheet["H2"] = GermanyStrings.RRPhone
sheet["I2"] = GermanyStrings.RRFax
sheet["J2"] = RRemail
sheet["K2"] = GermanyStrings.product
actorProfilesWorkbook.save(filename=actorProfiles)
actorProfilesWorkbook.close()
copy(actorProfiles, path)

# REMS Roster file
REMSRosterWorkbook = Workbook()
sheet = REMSRosterWorkbook.active
sheet["A1"] = "Territory_ID"
sheet["B1"] = "First_Name"
sheet["C1"] = "Last_Name"
sheet["D1"] = "Email"
sheet["E1"] = "Phone_Number"
sheet["F1"] = "Contact_Type"
sheet["G1"] = "Global_ID"

sheet["B2"] = GermanyStrings.FMRFirstName
sheet["C2"] = GermanyStrings.FMRLastName
sheet["D2"] = GermanyStrings.FMRTitle + str(globalID)[5:] + "@bms.com"
sheet["E2"] = GermanyStrings.FMRPhone
sheet["F2"] = GermanyStrings.FMRTitle
sheet["G2"] = globalID

sheet["B3"] = GermanyStrings.KAMFirstName
sheet["C3"] = GermanyStrings.KAMLastName
sheet["D3"] = GermanyStrings.KAMEmail
sheet["E3"] = GermanyStrings.KAMPhone
sheet["F3"] = GermanyStrings.KAMTitle
sheet["G3"] = globalID
REMSRosterWorkbook.save(filename=REMSRoster)
REMSRosterWorkbook.close()
copy(REMSRoster, path)

print("Files generated")
print("Files uploading")

cnopts = pysftp.CnOpts()
cnopts.hostkeys = None
with pysftp.Connection() as sftp:
    print("Connection established")
    actorProfileSource = path + "/" + actorProfiles
    destination = '/CART_NonProd/CART_VAL/MDM TO REMS/'
    with sftp.cd(remotepath=destination):
        files = sftp.listdir()
        for f in files:
            sftp.remove(f)
        sftp.put(path + "/" + actorProfiles)
        sftp.put(path + "/" + REMSRoster)
        sftp.put(path + "/" + siteMasterData)

print("Files uploaded")
print("Global ID: " + globalID)
print("RR Email: " + RRemail)
print("FMR Email: " + GermanyStrings.FMRTitle + str(globalID)[5:] + "@bms.com")
#
# #Time to run the admin job
#
# print("Running Chrome")
#
# options = Options()
# options.headless = True
# browser = webdriver.Chrome(options=options)
# browser.get('https://bms-cartrems-sit.pegacloud.net/prweb/')
#
# time.sleep(4)
#
# search = browser.find_element_by_id("txtUserID")
# search.send_keys(GermanyStrings.Admin)
# search = browser.find_element_by_id("txtPassword")
# search.send_keys(GermanyStrings.password)
# search.send_keys("\n")
#
# time.sleep(4)
# browser.find_element_by_tag_name('body').send_keys(Keys.CONTROL + 't')
#
# # Load a page
# browser.get("https://bms-cartrems-sit.pegacloud.net/prweb/5r-A_gl726GtgpdyVeQ3nw%28%28*/!TABTHREAD0?pyActivity=%40baseclass.pzTransformAndRun&pzTransactionId=&pzFromFrame=&pzPrimaryPageName=RunRecordWindow&preActivity=pzRunRecordInitializeWrapper&preActivityParams=%7B%22InsHandle%22%3A%22RULE-OBJ-ACTIVITY%20CELG-FW-CART-WORK%20PROCESSFROMKITEWORKS%20%2320201015T201100.059%20GMT%22%7D&action=display&harnessName=pxRunRecord&className=Pega-RunRecord&pyDataTransform=&pyPreActivity=doUIAction&pzPrimaryPage=RunRecordWindow&checkForNewPage=true&readOnly=false&frameName=RULE-OBJ-ACTIVITY%20CELG-FW-CART-WORK%20PROCESSFROMKITEWORKS%20%2320201015T201100.059%20GMT&target=popup&portalThreadName=STANDARD&portalName=Developer&pzHarnessID=HID86F49B111C504FD60AA267DD60D81C34")
#
# browser.find_element_by_xpath('//*[@id="RULE_KEY"]/div/div/div/div/div/div/button/div/div/div/div[contains(text(), "Run")]').click()
#
# print("processed finished")
#
# time.sleep(10)
# browser.quit()

