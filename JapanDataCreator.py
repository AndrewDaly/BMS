from openpyxl import Workbook
from openpyxl import load_workbook
from random import randint
from datetime import date
import os
import time
import string
import random
from shutil import copy
import JapanStrings
import pysftp

# Randomly generate a new global ID
lengthOfGlobalID = 9
globalID = ''.join(["{}".format(randint(0, 9)) for num in range(0, lengthOfGlobalID)])
print("Global ID: " + globalID)
lengthOfPhoneNumbers = 10
phoneNumber = "0" + ''.join(["{}".format(randint(0, 9)) for num in range(0, lengthOfPhoneNumbers)])

# Create a location to save created files
directory = "Japan Global ID " + globalID
path = os.path.join(os.getcwd(), directory)
os.mkdir(path)

# Customize RR email
RRemail = JapanStrings.RRFirstName + JapanStrings.RRLastName + str(globalID)[5:] + "@bms.com"
print("RR email: " + RRemail)
# Get timestamp for file naming
fdate = date.today().strftime('%m%d%Y')
timestamp = time.strftime("%H%M")
MMDDYYYYHHMM = str(fdate) + timestamp
actorProfiles = "Actor_Profile_Test_" + MMDDYYYYHHMM + ".xlsx"
REMSRoster = "Roster_Test_" + MMDDYYYYHHMM + ".xlsx"
siteMasterData = "SITE_MASTER_TEST_REMS_" + MMDDYYYYHHMM + ".xlsx"

# Create the Site Master Data Sheet
siteMasterDataworkbook = Workbook()
sheet = siteMasterDataworkbook.active
sheet["A1"] = "Global_ID"
sheet["B1"] = "AccountName"
sheet["C1"] = "AccountType"
sheet["D1"] = "Product_Activation"
sheet["E1"] = "DEA"
sheet["F1"] = "AddressLine1"
sheet["G1"] = "City"
sheet["H1"] = "State"
sheet["I1"] = "ZIP"
sheet["J1"] = "Country"
sheet["A2"] = globalID
sheet["B2"] = JapanStrings.accountName
sheet["C2"] = "Hospital"
sheet["D2"] = JapanStrings.product
sheet["F2"] = JapanStrings.address
sheet["G2"] = JapanStrings.City
sheet["H2"] = JapanStrings.prefecture
sheet["I2"] = JapanStrings.siteZip
sheet["J2"] = JapanStrings.country
siteMasterDataworkbook.save(filename=siteMasterData)
siteMasterDataworkbook.close()
copy(siteMasterData, path)

# actor profiles sheet
actorProfilesWorkbook = Workbook()
sheet = actorProfilesWorkbook.active
sheet["A1"] = "Role"
sheet["B1"] = "First Name"
sheet["C1"] = "Last Name"
sheet["D1"] = "Job Title"
sheet["E1"] = "Global_ID"
sheet["F1"] = "Account Name"
sheet["G1"] = "Credentials"
sheet["H1"] = "Primary Phone"
sheet["I1"] = "Fax"
sheet["J1"] = "Email Address"
sheet["K1"] = "Product_Activation"
sheet["A2"] = "Authorized Representative"
sheet["B2"] = JapanStrings.RRFirstName
sheet["C2"] = JapanStrings.RRLastName
sheet["D2"] = "Authorized Representative"
sheet["E2"] = globalID
sheet["F2"] = JapanStrings.accountName
sheet["G2"] = "MD"
sheet["H2"] = "0" + ''.join(["{}".format(randint(0, 9)) for num in range(0, lengthOfPhoneNumbers)])
sheet["I2"] = "0" + ''.join(["{}".format(randint(0, 9)) for num in range(0, lengthOfPhoneNumbers)])
sheet["J2"] = RRemail
sheet["K2"] = JapanStrings.product
actorProfilesWorkbook.save(filename=actorProfiles)
actorProfilesWorkbook.close()
copy(actorProfiles, path)

# REMS Roster file
REMSRosterWorkbook = Workbook()
sheet = REMSRosterWorkbook.active
sheet["A1"] = "Territory_ID"
sheet["B1"] = "First_Name"
sheet["C1"] = "Last_Name"
sheet["D1"] = "Email"
sheet["E1"] = "Phone_Number"
sheet["F1"] = "Contact_Type"
sheet["G1"] = "Global_ID"
sheet["B2"] = JapanStrings.MLFirstName
sheet["C2"] = JapanStrings.MLLastName
sheet["D2"] = JapanStrings.MLTitle + str(globalID)[5:] + "@bms.com"
print("ML Email: " + JapanStrings.MLTitle + str(globalID)[5:] + "@bms.com")
print("ML Password: Rules")
sheet["E2"] = "0" + ''.join(["{}".format(randint(0, 9)) for num in range(0, lengthOfPhoneNumbers)])
sheet["F2"] = JapanStrings.MLTitle
sheet["G2"] = globalID
sheet["B3"] = JapanStrings.SMFirstName
sheet["C3"] = JapanStrings.SMLastName
sheet["D3"] = JapanStrings.SMTitle + globalID + "@bms.com"
sheet["E3"] = "0" + ''.join(["{}".format(randint(0, 9)) for num in range(0, lengthOfPhoneNumbers)])
sheet["F3"] = JapanStrings.SMTitle
sheet["G3"] = globalID
REMSRosterWorkbook.save(filename=REMSRoster)
REMSRosterWorkbook.close()
copy(REMSRoster, path)

print("Files generated")


print("Files uploading")

cnopts = pysftp.CnOpts()
cnopts.hostkeys = None
with pysftp.Connection() as sftp:
    print("Connection established")
    actorProfileSource = path + "/" + actorProfiles
    destination = '/CART_NonProd/CART_VAL/MDM TO REMS/'
    with sftp.cd(remotepath=destination):
        files = sftp.listdir()
        for f in files:
            sftp.remove(f)
        sftp.put(path + "/" + actorProfiles)
        sftp.put(path + "/" + REMSRoster)
        sftp.put(path + "/" + siteMasterData)
print("Files uploaded")
