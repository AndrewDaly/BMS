from openpyxl import Workbook
from openpyxl import load_workbook
from random import randint
from datetime import date
import os
import time
import string
import random
from shutil import copy
import GeneralStrings
from selenium import webdriver
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.options import Options
import pysftp

country = input("Enter the country you want (Germany, Austria, Switzerland):")
print("You selected: " + country)
environment = input("Enter the environment you want (SIT, UAT):")
print("You selected: " + environment)

try:
    countryStrings = __import__(country+"Strings")
except ImportError:
    print("error importing")
    exit(8)

def insert_dot(string, index):
    return string[:index] + '.' + string[index:]

# Randomly generate a new global ID
lengthOfGlobalID = 9
globalID = ''.join(["{}".format(randint(0, 9)) for num in range(0, lengthOfGlobalID)])

# Create a location to save created files
directory = country + " Global ID " + globalID
path = os.path.join(os.getcwd(), directory)
os.mkdir(path)

# Randomize RR email
randomString = ''.join(random.choice(string.ascii_uppercase + string.digits) for _ in range(8))
RRemail = GeneralStrings.rminpRepresentativeEmail.replace("@gmail.com","") + "+" + randomString + "@gmail.com"


# Get timestamp for file naming
fdate = date.today().strftime('%m%d%Y')
timestamp = time.strftime("%H%M")
MMDDYYYYHHMM = str(fdate) + timestamp
actorProfiles = "Actor_Profile_Test_" + MMDDYYYYHHMM + ".xlsx"
REMSRoster = "Roster_Test_" + MMDDYYYYHHMM + ".xlsx"
siteMasterData = "SITE_MASTER_TEST_REMS_" + MMDDYYYYHHMM + ".xlsx"




# Create the Site Master Data Sheet
siteMasterDataworkbook = Workbook()
sheet = siteMasterDataworkbook.active
sheet["A1"] = "Global_ID"
sheet["B1"] = "AccountName"
sheet["C1"] = "AccountType"
sheet["D1"] = "Product_Activation"
sheet["E1"] = "DEA"
sheet["F1"] = "AddressLine1"
sheet["G1"] = "City"
sheet["H1"] = "State"
sheet["I1"] = "ZIP"
sheet["J1"] = "Country"
sheet["K1"] = "AffilEntityID"
sheet["L1"] = "AffilEntityName"
sheet["M1"] = "Affil_AccountType"
sheet["N1"] = "Affil_DEA"
sheet["O1"] = "Affil_AddressLine1"
sheet["P1"] = "Affil_City"
sheet["Q1"] = "Affil_State"
sheet["R1"] = "Affil_ZIP"
sheet["S1"] = "Affil_Country"
sheet["T1"] = "AffilEntityType"

sheet["A2"] = globalID
sheet["B2"] = GeneralStrings.accountName
sheet["C2"] = "Hospital"
sheet["D2"] = "Ide-Cel"
sheet["E2"] = ""
sheet["F2"] = countryStrings.siteAddress
sheet["G2"] = countryStrings.siteCity
sheet["H2"] = ""
sheet["I2"] = countryStrings.siteZip
sheet["J2"] = countryStrings.country
sheet["K2"] = ""
sheet["L2"] = ""
sheet["M2"] = ""
sheet["N2"] = ""
sheet["O2"] = ""
sheet["P2"] = ""
sheet["Q2"] = ""
sheet["R2"] = ""
sheet["S2"] = ""
sheet["T2"] = ""
siteMasterDataworkbook.save(filename=siteMasterData)
siteMasterDataworkbook.close()

copy(siteMasterData, path)

# actor profiles sheet
actorProfilesWorkbook = Workbook()
sheet = actorProfilesWorkbook.active
sheet["A1"] = "Role"
sheet["B1"] = "First Name"
sheet["C1"] = "Last Name"
sheet["D1"] = "Job Title"
sheet["E1"] = "Global_ID"
sheet["F1"] = "Account Name"
sheet["G1"] = "Credentials"
sheet["H1"] = "Primary Phone"
sheet["I1"] = "Fax"
sheet["J1"] = "Email Address"
sheet["K1"] = "Product_Activation"

sheet["A2"] = "Authorized Representative"
sheet["B2"] = GeneralStrings.RRFirstName
sheet["C2"] = GeneralStrings.RRLastName
sheet["D2"] = "Authorized Representative"
sheet["E2"] = globalID
sheet["F2"] = GeneralStrings.accountName
sheet["G2"] = countryStrings.RRCredentials
sheet["H2"] = countryStrings.RRPhone
sheet["I2"] = countryStrings.RRFax
sheet["J2"] = RRemail
sheet["K2"] = countryStrings.product
actorProfilesWorkbook.save(filename=actorProfiles)
actorProfilesWorkbook.close()
copy(actorProfiles, path)

# REMS Roster file
REMSRosterWorkbook = Workbook()
sheet = REMSRosterWorkbook.active
sheet["A1"] = "Territory_ID"
sheet["B1"] = "First_Name"
sheet["C1"] = "Last_Name"
sheet["D1"] = "Email"
sheet["E1"] = "Phone_Number"
sheet["F1"] = "Contact_Type"
sheet["G1"] = "Global_ID"

sheet["A2"] = ""
sheet["B2"] = GeneralStrings.FMRFirstName
sheet["C2"] = GeneralStrings.FMRLastName
sheet["D2"] = GeneralStrings.FMREmail
sheet["E2"] = countryStrings.FMRPhone
sheet["F2"] = countryStrings.FMRTitle
sheet["G2"] = globalID

sheet["A3"] = ""
sheet["B3"] = GeneralStrings.KAMFirstName
sheet["C3"] = GeneralStrings.KAMLastName
sheet["D3"] = GeneralStrings.KAMEmail
sheet["E3"] = countryStrings.KAMPhone
sheet["F3"] = countryStrings.KAMTitle
sheet["G3"] = globalID
REMSRosterWorkbook.save(filename=REMSRoster)
REMSRosterWorkbook.close()
copy(REMSRoster, path)


print("Files generated")
print("Files uploading")

cnopts = pysftp.CnOpts()
cnopts.hostkeys = None
with pysftp.Connection() as sftp:
    print("Connection established")
    actorProfileSource = path + "/" + actorProfiles
    destination = '/CART_NonProd/CART_SIT/MDM TO REMS/'
    with sftp.cd(remotepath=destination):

        files = sftp.listdir()

        for f in files:
            sftp.remove(f)

        sftp.put(path + "/" + actorProfiles)
        sftp.put(path + "/" + REMSRoster)
        sftp.put(path + "/" + siteMasterData)

print("Files uploaded")

#Time to run the admin job

print("Running Chrome")

options = Options()
options.headless = True
browser = webdriver.Chrome(options=options)
browser.get('https://bms-cartrems-sit.pegacloud.net/prweb/')

time.sleep(4)

search = browser.find_element_by_id("txtUserID")
search.send_keys(countryStrings.Admin)
search = browser.find_element_by_id("txtPassword")
search.send_keys(countryStrings.password)
search.send_keys("\n")

time.sleep(4)
browser.find_element_by_tag_name('body').send_keys(Keys.CONTROL + 't')

# Load a page
browser.get("https://bms-cartrems-sit.pegacloud.net/prweb/5r-A_gl726GtgpdyVeQ3nw%28%28*/!TABTHREAD0?pyActivity=%40baseclass.pzTransformAndRun&pzTransactionId=&pzFromFrame=&pzPrimaryPageName=RunRecordWindow&preActivity=pzRunRecordInitializeWrapper&preActivityParams=%7B%22InsHandle%22%3A%22RULE-OBJ-ACTIVITY%20CELG-FW-CART-WORK%20PROCESSFROMKITEWORKS%20%2320201015T201100.059%20GMT%22%7D&action=display&harnessName=pxRunRecord&className=Pega-RunRecord&pyDataTransform=&pyPreActivity=doUIAction&pzPrimaryPage=RunRecordWindow&checkForNewPage=true&readOnly=false&frameName=RULE-OBJ-ACTIVITY%20CELG-FW-CART-WORK%20PROCESSFROMKITEWORKS%20%2320201015T201100.059%20GMT&target=popup&portalThreadName=STANDARD&portalName=Developer&pzHarnessID=HID86F49B111C504FD60AA267DD60D81C34")

browser.find_element_by_xpath('//*[@id="RULE_KEY"]/div/div/div/div/div/div/button/div/div/div/div[contains(text(), "Run")]').click()

print("processed finished")

time.sleep(10)
browser.quit()



