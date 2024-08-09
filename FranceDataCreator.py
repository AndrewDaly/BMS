from openpyxl import Workbook
from openpyxl import load_workbook
from random import randint
from datetime import date
import os
import time
import string
import random
from shutil import copy
import FranceStrings
import pysftp

# Randomly generate a new global ID
lengthOfGlobalID = 9
globalID = ''.join(["{}".format(randint(0, 9)) for num in range(0, lengthOfGlobalID)])
print("Global ID: " + globalID)

# Create a location to save created files
directory = "France Global ID " + globalID
path = os.path.join(os.getcwd(), directory)
os.mkdir(path)

# Customize RR email
RRemail = FranceStrings.RRFirstName + FranceStrings.RRLastName + str(globalID)[5:] + "@bms.com"
# RRemail = "duboisaaron092@gmail.com"
print("RR email: " + RRemail)

# Get timestamp for file naming
fdate = date.today().strftime('%m%d%Y')
timestamp = time.strftime("%H%M")
MMDDYYYYHHMM = str(fdate) + timestamp
actorProfiles = "Actor_Profile_Test_" + MMDDYYYYHHMM + ".xlsx"
REMSRoster = "Roster_Test_" + MMDDYYYYHHMM + ".xlsx"
siteMasterData = "SITE_MASTER_TEST_REMS_" + MMDDYYYYHHMM + ".xlsx"

# Create the Site Master Data Sheet
siteMasterDataworkbook = Workbook()
sheet = siteMasterDataworkbook.active
sheet["A1"] = "Global_ID"
sheet["B1"] = "AccountName"
sheet["C1"] = "AccountType"
sheet["D1"] = "Product_Activation"
sheet["E1"] = "DEA"
sheet["F1"] = "AddressLine1"
sheet["G1"] = "City"
sheet["H1"] = "State"
sheet["I1"] = "ZIP"
sheet["J1"] = "Country"
sheet["K1"] = "AffilEntityID"
sheet["L1"] = "AffilEntityName"
sheet["M1"] = "Affil_AccountType"
sheet["N1"] = "Affil_DEA"
sheet["O1"] = "Affil_AddressLine1"
sheet["P1"] = "Affil_City"
sheet["Q1"] = "Affil_State"
sheet["R1"] = "Affil_ZIP"
sheet["S1"] = "Affil_Country"
sheet["T1"] = "AffilEntityType"
sheet["A2"] = globalID
sheet["B2"] = FranceStrings.accountName
sheet["C2"] = "Hospital"
sheet["D2"] = FranceStrings.product
sheet["F2"] = FranceStrings.siteAddress
sheet["G2"] = FranceStrings.siteCity
sheet["I2"] = FranceStrings.siteZip
sheet["J2"] = FranceStrings.country
siteMasterDataworkbook.save(filename=siteMasterData)
siteMasterDataworkbook.close()

copy(siteMasterData, path)

# actor profiles sheet
actorProfilesWorkbook = Workbook()
sheet = actorProfilesWorkbook.active
sheet["A1"] = "Role"
sheet["B1"] = "First Name"
sheet["C1"] = "Last Name"
sheet["D1"] = "Job Title"
sheet["E1"] = "Global_ID"
sheet["F1"] = "Account Name"
sheet["G1"] = "Credentials"
sheet["H1"] = "Primary Phone"
sheet["I1"] = "Fax"
sheet["J1"] = "Email Address"
sheet["K1"] = "Product_Activation"

sheet["A2"] = "Authorized Representative"
sheet["B2"] = FranceStrings.RRFirstName
sheet["C2"] = FranceStrings.RRLastName
sheet["D2"] = "Authorized Representative"
sheet["E2"] = globalID
sheet["F2"] = FranceStrings.accountName
sheet["G2"] = FranceStrings.RRCredentials
sheet["H2"] = FranceStrings.RRPhone
sheet["I2"] = FranceStrings.RRFax
sheet["J2"] = RRemail
sheet["K2"] = FranceStrings.product
actorProfilesWorkbook.save(filename=actorProfiles)
actorProfilesWorkbook.close()
copy(actorProfiles, path)

# REMS Roster file
REMSRosterWorkbook = Workbook()
sheet = REMSRosterWorkbook.active
sheet["A1"] = "Territory_ID"
sheet["B1"] = "First_Name"
sheet["C1"] = "Last_Name"
sheet["D1"] = "Email"
sheet["E1"] = "Phone_Number"
sheet["F1"] = "Contact_Type"
sheet["G1"] = "Global_ID"
sheet["B2"] = FranceStrings.FMRFirstName
sheet["C2"] = FranceStrings.FMRLastName
sheet["D2"] = FranceStrings.FMRTitle + str(globalID)[5:] + "@bms.com"
sheet["E2"] = FranceStrings.FMRPhone
sheet["F2"] = FranceStrings.FMRTitle #error from kiteworks email
sheet["G2"] = globalID
sheet["B3"] = FranceStrings.KAMFirstName
sheet["C3"] = FranceStrings.KAMLastName
sheet["D3"] = FranceStrings.KAMEmail
sheet["E3"] = FranceStrings.KAMPhone
sheet["F3"] = FranceStrings.KAMTitle #error from kiteworks email
sheet["G3"] = globalID
REMSRosterWorkbook.save(filename=REMSRoster)
REMSRosterWorkbook.close()
copy(REMSRoster, path)
print("Files generated")
print("Files uploading")
cnopts = pysftp.CnOpts()
cnopts.hostkeys = None
with pysftp.Connection() as sftp:
    print("Connection established")
    actorProfileSource = path + "/" + actorProfiles
    destination = '/CART_NonProd/CART_VAL/MDM TO REMS/'
    with sftp.cd(remotepath=destination):
        files = sftp.listdir()
        for f in files:
            sftp.remove(f)
        sftp.put(path + "/" + actorProfiles)
        sftp.put(path + "/" + REMSRoster)
        sftp.put(path + "/" + siteMasterData)
print("Files uploaded")