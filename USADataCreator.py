from openpyxl import Workbook
from openpyxl import load_workbook
from random import randint
from datetime import date
import os
import time
import string
import random
from shutil import copy
import USAStrings
import pysftp

# Randomly generate a new global ID
lengthOfGlobalID = 9
globalID = ''.join(["{}".format(randint(0, 9)) for num in range(0, lengthOfGlobalID)])

# Create a location to save created files
directory = "USA Global ID " + globalID
path = os.path.join(os.getcwd(), directory)
os.mkdir(path)

# Customize RR email
RRemail = USAStrings.RRFirstName + USAStrings.RRLastName + str(globalID)[5:] + "@bms.com"
#RRemail = USAStrings.RREmail
# Get timestamp for file naming
fdate = date.today().strftime('%m%d%Y')
timestamp = time.strftime("%H%M")
MMDDYYYYHHMM = str(fdate) + timestamp
actorProfiles = "Actor_Profile_Test_" + MMDDYYYYHHMM + ".xlsx"
REMSRoster = "REMS_Roster_Test_" + MMDDYYYYHHMM + ".xlsx"
siteMasterData = "SITE_MASTER_TEST_REMS_" + MMDDYYYYHHMM + ".xlsx"

# Create the Site Master Data Sheet
siteMasterDataworkbook = Workbook()
sheet = siteMasterDataworkbook.active
sheet["A1"] = "Global_ID"
sheet["B1"] = "AccountName"
sheet["C1"] = "AccountType"
sheet["D1"] = "Product_Activation"
sheet["E1"] = "DEA"
sheet["F1"] = "AddressLine1"
sheet["G1"] = "City"
sheet["H1"] = "State"
sheet["I1"] = "ZIP"
sheet["J1"] = "Country"
sheet["K1"] = "AffilEntityID"
sheet["L1"] = "AffilEntityName"
sheet["M1"] = "Affil_AccountType"
sheet["N1"] = "Affil_DEA"
sheet["O1"] = "Affil_AddressLine1"
sheet["P1"] = "Affil_City"
sheet["Q1"] = "Affil_State"
sheet["R1"] = "Affil_ZIP"
sheet["S1"] = "Affil_Country"
sheet["T1"] = "AffilEntityType"

sheet["A2"] = globalID
sheet["B2"] = USAStrings.accountName
sheet["C2"] = "Hospital"
sheet["D2"] = USAStrings.product
sheet["F2"] = USAStrings.address
sheet["G2"] = USAStrings.city
sheet["H2"] = USAStrings.state
sheet["I2"] = USAStrings.zip
sheet["J2"] = USAStrings.country

sheet["K2"] = str(int(globalID))
sheet["L2"] = USAStrings.AffilEntityName
sheet["M2"] = USAStrings.Affil_AccountType
sheet["O2"] = USAStrings.Affil_AddressLine1
sheet["P2"] = USAStrings.Affil_City
sheet["Q2"] = USAStrings.Affil_State
sheet["R2"] = USAStrings.Affil_ZIP
sheet["S2"] = USAStrings.Affil_Country
siteMasterDataworkbook.save(filename=siteMasterData)
siteMasterDataworkbook.close()
copy(siteMasterData, path)

# actor profiles sheet
actorProfilesWorkbook = Workbook()
sheet = actorProfilesWorkbook.active
sheet["A1"] = "Role"
sheet["B1"] = "First Name"
sheet["C1"] = "Last Name"
sheet["D1"] = "Job Title"
sheet["E1"] = "Global_ID"
sheet["F1"] = "Account Name"
sheet["G1"] = "Credentials"
sheet["H1"] = "Primary Phone"
sheet["I1"] = "Fax"
sheet["J1"] = "Email Address"
sheet["K1"] = "Product_Activation"

sheet["A2"] = "Authorized Representative"
sheet["B2"] = USAStrings.RRFirstName
sheet["C2"] = USAStrings.RRLastName
sheet["D2"] = "Authorized Representative"
sheet["E2"] = globalID
sheet["F2"] = USAStrings.accountName
sheet["G2"] = "MD"
sheet["H2"] = USAStrings.RRPhone
sheet["I2"] = USAStrings.RRFax
sheet["J2"] = RRemail
sheet["K2"] = USAStrings.product
actorProfilesWorkbook.save(filename=actorProfiles)
actorProfilesWorkbook.close()
copy(actorProfiles, path)

# REMS Roster file
REMSRosterWorkbook = Workbook()
sheet = REMSRosterWorkbook.active
sheet["A1"] = "Territory_ID"
sheet["B1"] = "First_Name"
sheet["C1"] = "Last_Name"
sheet["D1"] = "Email"
sheet["E1"] = "Phone_Number"
sheet["F1"] = "Contact_Type"
sheet["G1"] = "Global_ID"

sheet["B2"] = USAStrings.FMRFirstName
sheet["C2"] = USAStrings.FMRLastName
FMREmail = USAStrings.FMRTitle +  str(globalID)[5:] + "@bms.com"
sheet["D2"] = FMREmail
sheet["E2"] = USAStrings.FMRPhone
sheet["F2"] = USAStrings.FMRTitle
sheet["G2"] = globalID

sheet["B3"] = USAStrings.KAMFirstName
sheet["C3"] = USAStrings.KAMLastName
sheet["D3"] = USAStrings.KAMEmail
sheet["E3"] = USAStrings.KAMPhone
sheet["F3"] = USAStrings.KAMTitle
sheet["G3"] = globalID

sheet["B4"] = USAStrings.LEMFirstName
sheet["C4"] = USAStrings.LEMLastName
sheet["D4"] = USAStrings.LEMEmail
sheet["E4"] = USAStrings.LEMPhone
sheet["F4"] = USAStrings.LEMTitle
sheet["G4"] = globalID
REMSRosterWorkbook.save(filename=REMSRoster)
REMSRosterWorkbook.close()
copy(REMSRoster, path)
print("Global ID: " + globalID)
print("RR Email: " + RRemail)
print("FMR Email: " + FMREmail)
print("Files generated")

exit(8)
print("Files uploading")

cnopts = pysftp.CnOpts()
cnopts.hostkeys = None
with pysftp.Connection() as sftp:
    print("Connection established")
    actorProfileSource = path + "/" + actorProfiles
    destination = '/CART_NonProd/CART_VAL/MDM TO REMS/'
    with sftp.cd(remotepath=destination):
        files = sftp.listdir()
        for f in files:
            sftp.remove(f)
        sftp.put(path + "/" + actorProfiles)
        sftp.put(path + "/" + REMSRoster)
        sftp.put(path + "/" + siteMasterData)
print("Files uploaded")
#
# exit(8)
# #Time to run the admin job
#
# print("Running Chrome")
#
#
# options = Options()
# options.headless = False
# browser = webdriver.Chrome(options=options)
# browser.get('https://bms-cartrems-sit.pegacloud.net/prweb/')
#
# time.sleep(4)
#
# search = browser.find_element_by_id("txtUserID")
# search.send_keys(USAStrings.Admin)
# time.sleep(4)
# search = browser.find_element_by_id("txtPassword")
# search.send_keys(USAStrings.password)
# time.sleep(4)
# browser.find_element_by_xpath('//*[@id="sub"]/span').click()
# #search.send_keys("\n")
# time.sleep(4)
# browser.find_element_by_tag_name('body').send_keys(Keys.CONTROL + 't')
#
# # Load a page
# #browser.get("https://bms-cartrems-sit.pegacloud.net/prweb/5r-A_gl726GtgpdyVeQ3nw%28%28*/!TABTHREAD0?pyActivity=%40baseclass.pzTransformAndRun&pzTransactionId=&pzFromFrame=&pzPrimaryPageName=RunRecordWindow&preActivity=pzRunRecordInitializeWrapper&preActivityParams=%7B%22InsHandle%22%3A%22RULE-OBJ-ACTIVITY%20CELG-FW-CART-WORK%20PROCESSFROMKITEWORKS%20%2320201015T201100.059%20GMT%22%7D&action=display&harnessName=pxRunRecord&className=Pega-RunRecord&pyDataTransform=&pyPreActivity=doUIAction&pzPrimaryPage=RunRecordWindow&checkForNewPage=true&readOnly=false&frameName=RULE-OBJ-ACTIVITY%20CELG-FW-CART-WORK%20PROCESSFROMKITEWORKS%20%2320201015T201100.059%20GMT&target=popup&portalThreadName=STANDARD&portalName=Developer&pzHarnessID=HID86F49B111C504FD60AA267DD60D81C34")
# browser.get("https://bms-cartrems-uat.pegacloud.net/prweb/Bsz-O0fbsVqyjVPkK8anhayyhQSazH33fJTFC9TozNA*/!TABTHREAD0?pyActivity=%40baseclass.pzTransformAndRun&pzTransactionId=&pzFromFrame=&pzPrimaryPageName=RunRecordWindow&preActivity=pzRunRecordInitializeWrapper&preActivityParams=%7B%22InsHandle%22%3A%22RULE-OBJ-ACTIVITY%20CELG-FW-CART-WORK%20PROCESSFROMKITEWORKS%20%2320201015T201100.059%20GMT%22%7D&action=display&harnessName=pxRunRecord&className=Pega-RunRecord&pyDataTransform=&pyPreActivity=doUIAction&pzPrimaryPage=RunRecordWindow&checkForNewPage=true&readOnly=false&frameName=RULE-OBJ-ACTIVITY%20CELG-FW-CART-WORK%20PROCESSFROMKITEWORKS%20%2320201015T201100.059%20GMT&target=popup&portalThreadName=STANDARD&portalName=Developer&pzHarnessID=HID4CE55D414B3CBC37061EF21D45D5144B")
# #browser.find_element_by_xpath('//*[@id="RULE_KEY"]/div/div/div/div/div/div/button/div/div/div/div[contains(text(), "Run")]').click()
# browser.find_element_by_xpath('/html/body/div[3]/header/div/div/div/div/div/div/div/div/div/div/div[2]/div/div/div/div/div[2]/div/div/div/div/div/div/div/div/button/div/div/div/div/text()').click()
# print("processed finished")
